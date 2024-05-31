import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Load the workbook
wb = openpyxl.load_workbook('Book21.xlsx')

# Access the active sheet
ws = wb.active

# Define formatting options
# Define formatting options
highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


# Create a DifferentialStyle object and assign the PatternFill to it
highlight_style = DifferentialStyle(fill=highlight_fill)

# Define the rule to highlight duplicates
duplicate_rule = Rule(type="duplicateValues", dxf=highlight_style)
duplicate_rule.formula = ['1']  # Apply formatting to all duplicates

# Apply the rule to the desired range of cells
ws.conditional_formatting.add('B1:B222', duplicate_rule)

# Save the modified workbook
wb.save('your_modified_file.xlsx')
